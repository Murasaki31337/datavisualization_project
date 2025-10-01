import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import FormulaRule

load_dotenv()

conn = psycopg2.connect(
    dbname=os.getenv("PGDATABASE"),
    user=os.getenv("PGUSER"),
    password=os.getenv("PGPASSWORD"),
    host=os.getenv("PGHOST"),
    port=os.getenv("PGPORT")
)

# make folders if not exist
os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

# Helper function
def run_query(sql):
    return pd.read_sql(sql, conn)

# --- Pie chart: Team wins distribution (JOIN inside subquery) ---
df_wins = run_query("""
SELECT winner, wins
FROM (
    SELECT m.winner, COUNT(*) AS wins
    FROM matches m
    JOIN detailed_matches_maps dm
      ON m.match_id = dm.match_id
    GROUP BY m.winner
) t
ORDER BY wins DESC;
""")
plt.figure(figsize=(6,6))
plt.pie(df_wins["wins"], labels=df_wins["winner"], autopct='%1.1f%%')
plt.title("Team Wins Distribution")
plt.savefig("charts/pie_team_wins.png")
print(f"Pie chart saved. Rows: {len(df_wins)} — Distribution of wins by team (JOIN enriched).")

# --- Bar chart: Top 10 players by kills (INNER JOIN) ---
df_kills = run_query("""
SELECT p.player_name, SUM(p.kills) AS total_kills
FROM player_stats p
INNER JOIN detailed_matches_player_stats d
  ON p.player_name = d.player_name
GROUP BY p.player_name
ORDER BY total_kills DESC
LIMIT 10;
""")
df_kills.plot(kind="bar", x="player_name", y="total_kills", legend=False)
plt.ylabel("Kills")
plt.title("Top 10 Players by Kills")
plt.tight_layout()
plt.savefig("charts/bar_top_kills.png")
print(f"Bar chart saved. Rows: {len(df_kills)} — Top players by kills (with INNER JOIN).")

# --- Horizontal bar: Average ACS per team (simple, keep clean) ---
df_acs = run_query("""
SELECT team, ROUND(AVG(acs)::numeric,1) AS avg_acs
FROM player_stats
GROUP BY team
ORDER BY avg_acs DESC
LIMIT 10;
""")
df_acs.plot(kind="barh", x="team", y="avg_acs", legend=False)
plt.xlabel("Average ACS")
plt.title("Average ACS per Team")
plt.tight_layout()
plt.savefig("charts/hbar_avg_acs.png")
print(f"Horizontal bar saved. Rows: {len(df_acs)} — ACS by team.")

# --- Line chart: Average player rating over time (LEFT JOIN with maps) ---
df_line = run_query("""
SELECT d.match_date, m.map_name, ROUND(AVG(d.rating)::numeric,2) AS avg_rating
FROM detailed_matches_player_stats d
LEFT JOIN detailed_matches_maps m
  ON d.match_id = m.match_id
WHERE d.rating IS NOT NULL
GROUP BY d.match_date, m.map_name
ORDER BY d.match_date;
""")
df_line.plot(kind="line", x="match_date", y="avg_rating", marker="o", color="purple")
plt.title("Average Player Rating Over Time (per map context)")
plt.ylabel("Avg Rating")
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("charts/line_avg_rating.png")
print(f"Line chart saved. Rows: {len(df_line)} — Average rating over time (LEFT JOIN with maps).")

# --- Histogram: Distribution of player ratings (JOIN with maps) ---
df_ratings = run_query("""
SELECT d.rating, m.map_name
FROM detailed_matches_player_stats d
JOIN detailed_matches_maps m
  ON d.match_id = m.match_id
WHERE d.rating IS NOT NULL;
""")
plt.figure()
plt.hist(df_ratings["rating"], bins=20, color="skyblue", edgecolor="black")
plt.title("Distribution of Player Ratings (by map)")
plt.xlabel("Rating")
plt.ylabel("Frequency")
plt.tight_layout()
plt.savefig("charts/hist_ratings.png")
print(f"Histogram saved. Rows: {len(df_ratings)} — Rating distribution (JOIN with maps).")

# --- Scatter plot: ADR vs ACS (JOIN with maps for context) ---
df_scatter = run_query("""
SELECT d.player_team, d.acs, d.adr, m.map_name
FROM detailed_matches_player_stats d
JOIN detailed_matches_maps m
  ON d.match_id = m.match_id
WHERE d.acs IS NOT NULL AND d.adr IS NOT NULL
LIMIT 300;
""")
plt.figure()
plt.scatter(df_scatter["acs"], df_scatter["adr"], alpha=0.6)
plt.xlabel("ACS")
plt.ylabel("ADR")
plt.title("ACS vs ADR (sample, JOIN with maps)")
plt.tight_layout()
plt.savefig("charts/scatter_acs_vs_adr.png")
print(f"Scatter plot saved. Rows: {len(df_scatter)} — Relation between ACS & ADR (JOIN with maps).")

# --- Plotly interactive line with range selector (multiple teams over time) ---
df_slider = run_query("""
SELECT match_date, player_team, ROUND(AVG(acs)::numeric,1) AS avg_acs
FROM detailed_matches_player_stats
WHERE acs IS NOT NULL
GROUP BY match_date, player_team
ORDER BY match_date;
""")

fig = px.line(
    df_slider,
    x="match_date", y="avg_acs",
    color="player_team",
    title="Team ACS Over Time (Interactive Lines with Range Selector)"
)

fig.update_layout(
    xaxis=dict(
        rangeselector=dict(
            buttons=list([
                dict(count=1, label="1d", step="day", stepmode="backward"),
                dict(count=7, label="1w", step="day", stepmode="backward"),
                dict(count=1, label="1m", step="month", stepmode="backward"),
                dict(step="all", label="All")
            ])
        ),
        rangeslider=dict(visible=True),
        type="date"
    ),
    yaxis_title="Average ACS",
    legend_title="Team"
)

fig.write_html("charts/time_slider.html")
print("✅ Interactive line chart with range selector saved as charts/time_slider.html")

# --- Export to Excel with formatting ---
export_data = {
    "TopPlayers": df_kills,
    "TeamWins": df_wins,
    "RatingTimeline": df_line
}
out_file = "exports/report.xlsx"

with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
    for sheet, df in export_data.items():
        df.to_excel(writer, sheet_name=sheet, index=False)



from openpyxl.formatting.rule import FormulaRule

# Apply formatting
wb = load_workbook(out_file)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    
    # Only apply conditional formatting to RatingTimeline
    if sheet == "RatingTimeline":
        for col in ws.iter_cols(min_row=2):
            # Apply only to numeric columns
            if all(isinstance(cell.value, (int, float, type(None))) for cell in col):
                col_letter = col[0].column_letter
                start_row = col[1].row
                end_row = col[-1].row
                rng = f"{col_letter}{start_row}:{col_letter}{end_row}"

                # Above average → green
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f"{col_letter}{start_row}>AVERAGE({rng})"],
                        fill=PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
                    )
                )
                # Below average → red
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f"{col_letter}{start_row}<AVERAGE({rng})"],
                        fill=PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    )
                )

wb.save(out_file)
print(f"✅ Excel report created with conditional formatting only in RatingTimeline: {out_file}, {len(export_data)} sheets.")
